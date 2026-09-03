import io
import json
import os
import sqlite3
import secrets
import urllib.error
import urllib.request
from datetime import date, datetime, timedelta
from functools import wraps
from pathlib import Path

from flask import Flask, Response, g, jsonify, request, send_from_directory, session
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import check_password_hash, generate_password_hash

try:
    import email_config
except ModuleNotFoundError:
    # Absent en production par design (voir .gitignore) : les réglages viennent
    # alors uniquement des variables d'environnement, via parametre_email().
    email_config = None

BASE_DIR = Path(__file__).resolve().parent

# En production, DATA_DIR pointe vers un volume persistant (ex: /data sur Railway) pour que
# la base de données survive aux redéploiements. En local, elle reste à côté du code.
DATA_DIR = Path(os.environ.get("DATA_DIR", BASE_DIR))
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = DATA_DIR / "app.db"
SECRET_KEY_PATH = BASE_DIR / "secret.key"

app = Flask(__name__, static_folder="static", static_url_path="")
# Railway (et la plupart des hébergeurs) placent l'app derrière un proxy : sans ça,
# request.remote_addr renverrait l'IP interne du proxy pour tout le monde, rendant
# la limitation par IP inutile.
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1)

# En production, définissez la variable d'environnement SECRET_KEY (l'hébergeur la fournit
# dans son tableau de bord). En local, une clé est générée une fois et stockée dans secret.key.
if os.environ.get("SECRET_KEY"):
    app.secret_key = os.environ["SECRET_KEY"]
else:
    if not SECRET_KEY_PATH.exists():
        SECRET_KEY_PATH.write_text(secrets.token_hex(32))
    app.secret_key = SECRET_KEY_PATH.read_text().strip()
app.config.update(SESSION_COOKIE_HTTPONLY=True, SESSION_COOKIE_SAMESITE="Lax")


def parametre_email(nom):
    """Variable d'environnement en priorité (production), sinon email_config.py (local)."""
    valeur_locale = getattr(email_config, nom, "") if email_config else ""
    return os.environ.get(nom) or valeur_locale


# ---------- Base de données ----------

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
        g.db.execute("PRAGMA journal_mode = WAL")
        g.db.execute("PRAGMA busy_timeout = 5000")
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


SCHEMA = """
CREATE TABLE IF NOT EXISTS entreprises (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nom TEXT NOT NULL,
    couleur TEXT NOT NULL,
    objectif_ca_ht_brut REAL
);

CREATE TABLE IF NOT EXISTS commerciaux (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nom TEXT NOT NULL,
    entreprise_id INTEGER NOT NULL REFERENCES entreprises(id),
    taux REAL,
    actif INTEGER NOT NULL DEFAULT 1,
    objectif_ca_ht_brut REAL
);

CREATE TABLE IF NOT EXISTS utilisateurs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    identifiant TEXT UNIQUE NOT NULL,
    password_hash TEXT NOT NULL,
    role TEXT NOT NULL CHECK(role IN ('admin', 'admin_secondaire', 'commercial', 'animateur')),
    commercial_id INTEGER REFERENCES commerciaux(id),
    nom TEXT,
    kpi_order TEXT,
    kpi_reduits TEXT
);

CREATE TABLE IF NOT EXISTS dossiers (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    commercial_id INTEGER NOT NULL REFERENCES commerciaux(id),
    date TEXT NOT NULL,
    client TEXT NOT NULL,
    voiture TEXT NOT NULL,
    plaque TEXT,
    garantie_achat REAL NOT NULL DEFAULT 0,
    garantie_prix_vendu REAL NOT NULL DEFAULT 0,
    mandat_total REAL NOT NULL DEFAULT 0,
    frais_intermediation REAL NOT NULL DEFAULT 0,
    nettoyage REAL NOT NULL DEFAULT 0,
    commission_agence REAL NOT NULL DEFAULT 0,
    achat_livraison REAL NOT NULL DEFAULT 0,
    vente_livraison REAL NOT NULL DEFAULT 0,
    exporte_excel_le TEXT,
    role_commercial TEXT NOT NULL DEFAULT 'les_deux'
);

CREATE TABLE IF NOT EXISTS retrocommissions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    dossier_id INTEGER NOT NULL UNIQUE REFERENCES dossiers(id),
    commercial_id INTEGER NOT NULL REFERENCES commerciaux(id),
    role TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS reinitialisations_mdp (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    utilisateur_id INTEGER NOT NULL REFERENCES utilisateurs(id),
    code_hash TEXT NOT NULL,
    expire_le TEXT NOT NULL,
    utilise INTEGER NOT NULL DEFAULT 0
);

CREATE TABLE IF NOT EXISTS tentatives (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    action TEXT NOT NULL,
    identifiant TEXT,
    ip TEXT NOT NULL,
    reussi INTEGER NOT NULL,
    horodatage TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS avis_google (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    dossier_id INTEGER NOT NULL REFERENCES dossiers(id),
    commercial_id INTEGER NOT NULL REFERENCES commerciaux(id),
    nom_client TEXT NOT NULL,
    role TEXT NOT NULL CHECK(role IN ('vendeur', 'acheteur')),
    cree_le TEXT NOT NULL,
    UNIQUE(dossier_id, role)
);

-- Ancien compteur mensuel d'avis Google (avant les fiches individuelles), conservé en
-- lecture seule pour les mois déjà écoulés au moment de la migration (voir migrer_db
-- et avis_obtenus_mensuel). Créée vide dès le premier démarrage sur les installations
-- neuves, pour que le code n'ait jamais à distinguer "table absente" de "table vide".
CREATE TABLE IF NOT EXISTS avis_google_legacy (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    commercial_id INTEGER NOT NULL REFERENCES commerciaux(id),
    mois TEXT NOT NULL,
    avis_obtenus INTEGER NOT NULL DEFAULT 0,
    UNIQUE(commercial_id, mois)
);

CREATE TABLE IF NOT EXISTS refacturations (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    commercial_id INTEGER NOT NULL REFERENCES commerciaux(id),
    mois TEXT NOT NULL,
    nombre_call_center INTEGER NOT NULL DEFAULT 0,
    nombre_leads_meta INTEGER NOT NULL DEFAULT 0,
    montant_leboncoin REAL NOT NULL DEFAULT 0,
    UNIQUE(commercial_id, mois)
);

CREATE TABLE IF NOT EXISTS tarifs_refacturation (
    categorie TEXT PRIMARY KEY,
    prix_unitaire REAL NOT NULL DEFAULT 0
);

CREATE TABLE IF NOT EXISTS charges_fixes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    entreprise_id INTEGER NOT NULL REFERENCES entreprises(id),
    mois TEXT NOT NULL,
    nom TEXT NOT NULL,
    montant REAL NOT NULL DEFAULT 0
);
"""


def migrer_db(db):
    colonnes = {row["name"] for row in db.execute("PRAGMA table_info(dossiers)")}
    if "frais_intermediation" not in colonnes:
        db.execute("ALTER TABLE dossiers ADD COLUMN frais_intermediation REAL NOT NULL DEFAULT 0")
    if "exporte_excel_le" not in colonnes:
        db.execute("ALTER TABLE dossiers ADD COLUMN exporte_excel_le TEXT")
    if "nettoyage" not in colonnes:
        db.execute("ALTER TABLE dossiers ADD COLUMN nettoyage REAL NOT NULL DEFAULT 0")
    if "commission_agence" not in colonnes:
        db.execute("ALTER TABLE dossiers ADD COLUMN commission_agence REAL NOT NULL DEFAULT 0")
    if "achat_livraison" not in colonnes:
        db.execute("ALTER TABLE dossiers ADD COLUMN achat_livraison REAL NOT NULL DEFAULT 0")
    if "vente_livraison" not in colonnes:
        db.execute("ALTER TABLE dossiers ADD COLUMN vente_livraison REAL NOT NULL DEFAULT 0")
    if "role_commercial" not in colonnes:
        db.execute("ALTER TABLE dossiers ADD COLUMN role_commercial TEXT NOT NULL DEFAULT 'les_deux'")

    colonnes_commerciaux = {row["name"] for row in db.execute("PRAGMA table_info(commerciaux)")}
    if "objectif_ca_ht_brut" not in colonnes_commerciaux:
        db.execute("ALTER TABLE commerciaux ADD COLUMN objectif_ca_ht_brut REAL")
    if "supprime" not in colonnes_commerciaux:
        db.execute("ALTER TABLE commerciaux ADD COLUMN supprime INTEGER NOT NULL DEFAULT 0")

    colonnes_entreprises = {row["name"] for row in db.execute("PRAGMA table_info(entreprises)")}
    if "objectif_ca_ht_brut" not in colonnes_entreprises:
        db.execute("ALTER TABLE entreprises ADD COLUMN objectif_ca_ht_brut REAL")

    colonnes_utilisateurs = {row["name"] for row in db.execute("PRAGMA table_info(utilisateurs)")}
    if "nom" not in colonnes_utilisateurs:
        db.execute("ALTER TABLE utilisateurs ADD COLUMN nom TEXT")
    if "kpi_order" not in colonnes_utilisateurs:
        db.execute("ALTER TABLE utilisateurs ADD COLUMN kpi_order TEXT")
    if "kpi_reduits" not in colonnes_utilisateurs:
        db.execute("ALTER TABLE utilisateurs ADD COLUMN kpi_reduits TEXT")

    schema_utilisateurs = db.execute(
        "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'utilisateurs'"
    ).fetchone()[0]
    if "'animateur'" not in schema_utilisateurs:
        db.executescript(
            """
            CREATE TABLE utilisateurs_nouveau (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                identifiant TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL CHECK(role IN ('admin', 'commercial', 'animateur')),
                commercial_id INTEGER REFERENCES commerciaux(id),
                nom TEXT,
                kpi_order TEXT,
                kpi_reduits TEXT
            );
            INSERT INTO utilisateurs_nouveau (id, identifiant, password_hash, role, commercial_id, nom, kpi_order, kpi_reduits)
                SELECT id, identifiant, password_hash, role, commercial_id, nom, kpi_order, kpi_reduits FROM utilisateurs;
            DROP TABLE utilisateurs;
            ALTER TABLE utilisateurs_nouveau RENAME TO utilisateurs;
            """
        )

    schema_utilisateurs = db.execute(
        "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'utilisateurs'"
    ).fetchone()[0]
    if "'admin_secondaire'" not in schema_utilisateurs:
        db.executescript(
            """
            CREATE TABLE utilisateurs_nouveau (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                identifiant TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL CHECK(role IN ('admin', 'admin_secondaire', 'commercial', 'animateur')),
                commercial_id INTEGER REFERENCES commerciaux(id),
                nom TEXT,
                kpi_order TEXT,
                kpi_reduits TEXT
            );
            INSERT INTO utilisateurs_nouveau (id, identifiant, password_hash, role, commercial_id, nom, kpi_order, kpi_reduits)
                SELECT id, identifiant, password_hash, role, commercial_id, nom, kpi_order, kpi_reduits FROM utilisateurs;
            DROP TABLE utilisateurs;
            ALTER TABLE utilisateurs_nouveau RENAME TO utilisateurs;
            """
        )

    # Charges fixes : passage d'une liste globale unique à une liste par entreprise et
    # par mois. Les anciennes lignes (sans notion d'entreprise/mois) sont rattachées à
    # la première entreprise existante et au mois courant plutôt que d'être perdues.
    colonnes_charges = {row["name"] for row in db.execute("PRAGMA table_info(charges_fixes)")}
    if colonnes_charges and "entreprise_id" not in colonnes_charges:
        anciennes_charges = db.execute("SELECT nom, montant FROM charges_fixes").fetchall()
        premiere_entreprise = db.execute("SELECT id FROM entreprises ORDER BY id LIMIT 1").fetchone()
        mois_courant = datetime.utcnow().strftime("%Y-%m")
        db.execute("DROP TABLE charges_fixes")
        db.execute(
            """CREATE TABLE charges_fixes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                entreprise_id INTEGER NOT NULL REFERENCES entreprises(id),
                mois TEXT NOT NULL,
                nom TEXT NOT NULL,
                montant REAL NOT NULL DEFAULT 0
            )"""
        )
        if premiere_entreprise is not None:
            for ancienne in anciennes_charges:
                db.execute(
                    "INSERT INTO charges_fixes (entreprise_id, mois, nom, montant) VALUES (?, ?, ?, ?)",
                    (premiere_entreprise["id"], mois_courant, ancienne["nom"], ancienne["montant"]),
                )

    # Avis Google : passage d'un simple compteur mensuel saisi par l'admin à des fiches
    # individuelles (nom du client, vendeur/acheteur, dossier lié) ajoutées par le
    # commercial ou l'admin. L'ancienne table ne contenait qu'un nombre — sans nom ni
    # dossier associé, impossible à convertir en fiche détaillée — mais elle contenait
    # de vraies données (pas de simples zéros de test) : on la renomme donc en table
    # "legacy" conservée telle quelle plutôt que de la supprimer, pour que les mois déjà
    # écoulés gardent exactement le nombre d'avis déjà utilisé dans leurs commissions.
    # Seuls les mois à partir de maintenant utilisent la nouvelle table structurée (voir
    # avis_obtenus_mensuel) — le mois en cours au moment de la migration doit donc être
    # ressaisi via la nouvelle interface s'il contenait déjà des avis.
    colonnes_avis = {row["name"] for row in db.execute("PRAGMA table_info(avis_google)")}
    if colonnes_avis and "dossier_id" not in colonnes_avis:
        # avis_google_legacy existe déjà (vide) depuis le SCHEMA exécuté juste avant :
        # on la supprime pour faire place à l'ancienne table, qui contient les vraies
        # données à conserver.
        db.execute("DROP TABLE IF EXISTS avis_google_legacy")
        db.execute("ALTER TABLE avis_google RENAME TO avis_google_legacy")
        db.execute(
            """CREATE TABLE avis_google (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                dossier_id INTEGER NOT NULL REFERENCES dossiers(id),
                commercial_id INTEGER NOT NULL REFERENCES commerciaux(id),
                nom_client TEXT NOT NULL,
                role TEXT NOT NULL CHECK(role IN ('vendeur', 'acheteur')),
                cree_le TEXT NOT NULL,
                UNIQUE(dossier_id, role)
            )"""
        )

    schema_commerciaux = db.execute(
        "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'commerciaux'"
    ).fetchone()[0]
    if "taux REAL NOT NULL" in schema_commerciaux:
        db.executescript(
            """
            CREATE TABLE commerciaux_nouveau (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nom TEXT NOT NULL,
                entreprise_id INTEGER NOT NULL REFERENCES entreprises(id),
                taux REAL,
                actif INTEGER NOT NULL DEFAULT 1
            );
            INSERT INTO commerciaux_nouveau (id, nom, entreprise_id, taux, actif)
                SELECT id, nom, entreprise_id, taux, actif FROM commerciaux;
            DROP TABLE commerciaux;
            ALTER TABLE commerciaux_nouveau RENAME TO commerciaux;
            """
        )


def init_db():
    db = sqlite3.connect(DB_PATH, timeout=30)
    db.row_factory = sqlite3.Row
    db.executescript(SCHEMA)
    db.commit()
    # BEGIN IMMEDIATE prend le verrou d'écriture dès le début de la transaction : gunicorn
    # démarre plusieurs workers en même temps et chacun appelle init_db() sur la même base —
    # sans ça, chacun vérifierait "colonne/compte absent ?" en parallèle avant qu'aucun n'ait
    # écrit, et tenterait la même migration ou le même seed en double (ALTER TABLE en double →
    # "duplicate column name", ou double création du compte admin → UNIQUE constraint failed).
    # Ici, tout ce qui est check-then-write est regroupé sous un seul verrou, donc sérialisé.
    db.execute("BEGIN IMMEDIATE")
    migrer_db(db)
    for categorie in ("call_center", "leads_meta"):
        db.execute(
            "INSERT OR IGNORE INTO tarifs_refacturation (categorie, prix_unitaire) VALUES (?, 0)",
            (categorie,),
        )
    already_seeded = db.execute("SELECT COUNT(*) FROM entreprises").fetchone()[0] > 0
    if not already_seeded:
        seed(db)
    db.commit()
    db.close()


def seed(db):
    entreprises = [
        ("Norda Bâtiment", "#E8A33D"),
        ("Lumis Énergie", "#4FD1C5"),
        ("Verrel Distribution", "#C58AF0"),
    ]
    db.executemany("INSERT INTO entreprises (nom, couleur) VALUES (?, ?)", entreprises)
    ent_ids = [r[0] for r in db.execute("SELECT id FROM entreprises ORDER BY id")]

    commerciaux = [
        ("Awa Diallo", ent_ids[0], 0.08, "awa.diallo"),
        ("Marc Lefèvre", ent_ids[0], 0.07, "marc.lefevre"),
        ("Julie Rambert", ent_ids[1], 0.10, "julie.rambert"),
        ("Karim Belhadj", ent_ids[2], 0.06, "karim.belhadj"),
    ]
    commercial_ids = {}
    for nom, ent_id, taux, identifiant in commerciaux:
        cur = db.execute(
            "INSERT INTO commerciaux (nom, entreprise_id, taux) VALUES (?, ?, ?)",
            (nom, ent_id, taux),
        )
        commercial_ids[identifiant] = cur.lastrowid
        db.execute(
            "INSERT INTO utilisateurs (identifiant, password_hash, role, commercial_id) VALUES (?, ?, 'commercial', ?)",
            (identifiant, generate_password_hash("commercial123", method="pbkdf2:sha256"), cur.lastrowid),
        )

    db.execute(
        "INSERT INTO utilisateurs (identifiant, password_hash, role, commercial_id) VALUES (?, ?, 'admin', NULL)",
        ("admin", generate_password_hash("admin123", method="pbkdf2:sha256")),
    )

    ventes = [
        (commercial_ids["awa.diallo"], "2026-07-03", "M. Beaumont", "Peugeot 3008", "EF-482-GT", 15200, 18400, 1200),
        (commercial_ids["awa.diallo"], "2026-07-14", "Ateliers du Nord", "Renault Trafic", "AB-113-CD", 8100, 9200, 650),
        (commercial_ids["marc.lefevre"], "2026-07-08", "Mme Ferran", "BMW Série 3", "GH-905-JK", 23800, 26500, 1800),
        (commercial_ids["julie.rambert"], "2026-07-11", "EcoVolt Sarl", "Tesla Model 3", "LM-227-NO", 37500, 41200, 2400),
        (commercial_ids["julie.rambert"], "2026-07-22", "Résidence Les Tilleuls", "Citroën C3", "PQ-664-RS", 13900, 15800, 900),
        (commercial_ids["karim.belhadj"], "2026-07-18", "Distri Ouest", "Ford Transit", "TU-338-VW", 10800, 12300, 780),
    ]
    db.executemany(
        """INSERT INTO dossiers
           (commercial_id, date, client, voiture, plaque, garantie_achat, garantie_prix_vendu, mandat_total)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        ventes,
    )


# ---------- Protection contre les tentatives répétées ----------

FENETRE_TENTATIVES_MINUTES = 15


def trop_de_tentatives(db, action, identifiant, seuil_identifiant, seuil_ip):
    limite = (datetime.utcnow() - timedelta(minutes=FENETRE_TENTATIVES_MINUTES)).isoformat()
    ip = request.remote_addr or "inconnue"

    if identifiant:
        n = db.execute(
            "SELECT COUNT(*) FROM tentatives WHERE action = ? AND identifiant = ? AND reussi = 0 AND horodatage > ?",
            (action, identifiant, limite),
        ).fetchone()[0]
        if n >= seuil_identifiant:
            return True

    n_ip = db.execute(
        "SELECT COUNT(*) FROM tentatives WHERE action = ? AND ip = ? AND reussi = 0 AND horodatage > ?",
        (action, ip, limite),
    ).fetchone()[0]
    return n_ip >= seuil_ip


def enregistrer_tentative(db, action, identifiant, reussi):
    ip = request.remote_addr or "inconnue"
    db.execute(
        "INSERT INTO tentatives (action, identifiant, ip, reussi, horodatage) VALUES (?, ?, ?, ?, ?)",
        (action, identifiant, ip, 1 if reussi else 0, datetime.utcnow().isoformat()),
    )
    # Ménage discret : évite que la table ne grossisse indéfiniment.
    limite_nettoyage = (datetime.utcnow() - timedelta(hours=24)).isoformat()
    db.execute("DELETE FROM tentatives WHERE horodatage < ?", (limite_nettoyage,))


def reponse_trop_de_tentatives():
    return jsonify(error="Trop de tentatives. Réessayez dans quelques minutes."), 429


# ---------- Auth helpers ----------

def current_user():
    uid = session.get("uid")
    if not uid:
        return None
    db = get_db()
    return db.execute("SELECT * FROM utilisateurs WHERE id = ?", (uid,)).fetchone()


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = current_user()
        if user is None:
            return jsonify(error="Non authentifié"), 401
        g.user = user
        return fn(*args, **kwargs)
    return wrapper


def est_type_admin(user):
    """Admin principal ou admin secondaire : les deux gèrent dossiers, équipe et
    entreprises à l'identique. Seuls les tarifs de refacturation et les charges
    fixes restent réservés à l'admin principal (voir admin_complet_requis)."""
    return user["role"] in ("admin", "admin_secondaire")


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = current_user()
        if user is None:
            return jsonify(error="Non authentifié"), 401
        if not est_type_admin(user):
            return jsonify(error="Réservé aux administrateurs"), 403
        g.user = user
        return fn(*args, **kwargs)
    return wrapper


def admin_complet_requis(fn):
    """Comme admin_required, mais exclut l'admin secondaire : réservé aux actions
    qui lui sont explicitement interdites (tarifs de refacturation, charges fixes)."""
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = current_user()
        if user is None:
            return jsonify(error="Non authentifié"), 401
        if user["role"] != "admin":
            return jsonify(error="Réservé à l'administrateur principal"), 403
        g.user = user
        return fn(*args, **kwargs)
    return wrapper


def ecriture_requise(fn):
    """Comme login_required, mais bloque le rôle animateur (accès lecture seule)."""
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = current_user()
        if user is None:
            return jsonify(error="Non authentifié"), 401
        if user["role"] == "animateur":
            return jsonify(error="Accès en lecture seule"), 403
        g.user = user
        return fn(*args, **kwargs)
    return wrapper


def peut_tout_voir(user):
    return user["role"] in ("admin", "admin_secondaire", "animateur")


def user_public(user, db):
    # L'ordre et l'état réduit des cartes KPI sont des préférences personnelles :
    # chaque compte (admin, commercial, animateur) a les siennes, stockées sur sa
    # propre ligne.
    kpi_order = json.loads(user["kpi_order"]) if user["kpi_order"] else None
    kpi_reduits = json.loads(user["kpi_reduits"]) if user["kpi_reduits"] else []

    if user["role"] == "admin":
        return {
            "id": user["id"],
            "role": "admin",
            "nom": user["nom"] or "Administrateur",
            "identifiant": user["identifiant"],
            "kpi_order": kpi_order,
            "kpi_reduits": kpi_reduits,
        }
    if user["role"] == "admin_secondaire":
        return {
            "id": user["id"],
            "role": "admin_secondaire",
            "nom": user["nom"] or "Administrateur secondaire",
            "identifiant": user["identifiant"],
            "kpi_order": kpi_order,
            "kpi_reduits": kpi_reduits,
        }
    if user["role"] == "animateur":
        return {
            "id": user["id"],
            "role": "animateur",
            "nom": user["nom"] or "Animateur",
            "identifiant": user["identifiant"],
            "kpi_order": kpi_order,
            "kpi_reduits": kpi_reduits,
        }
    com = db.execute("SELECT * FROM commerciaux WHERE id = ?", (user["commercial_id"],)).fetchone()
    return {
        "id": user["id"],
        "role": "commercial",
        "nom": com["nom"],
        "identifiant": user["identifiant"],
        "commercial_id": com["id"],
        "entreprise_id": com["entreprise_id"],
        "taux": com["taux"],
        "kpi_order": kpi_order,
        "kpi_reduits": kpi_reduits,
    }


# ---------- Routes: auth ----------

@app.post("/api/login")
def login():
    data = request.get_json(silent=True) or {}
    identifiant = (data.get("identifiant") or "").strip()
    mot_de_passe = data.get("mot_de_passe") or ""
    db = get_db()

    if trop_de_tentatives(db, "login", identifiant, seuil_identifiant=5, seuil_ip=20):
        return reponse_trop_de_tentatives()

    user = db.execute("SELECT * FROM utilisateurs WHERE identifiant = ?", (identifiant,)).fetchone()
    if user is None or not check_password_hash(user["password_hash"], mot_de_passe):
        enregistrer_tentative(db, "login", identifiant, reussi=False)
        db.commit()
        return jsonify(error="Identifiant ou mot de passe incorrect"), 401
    if user["role"] == "commercial":
        com = db.execute("SELECT actif FROM commerciaux WHERE id = ?", (user["commercial_id"],)).fetchone()
        if com is None or not com["actif"]:
            enregistrer_tentative(db, "login", identifiant, reussi=False)
            db.commit()
            return jsonify(error="Ce compte est désactivé"), 403

    enregistrer_tentative(db, "login", identifiant, reussi=True)
    db.commit()
    session.clear()
    session["uid"] = user["id"]
    session.permanent = True
    return jsonify(user_public(user, db))


@app.post("/api/logout")
def logout():
    session.clear()
    return jsonify(ok=True)


def envoyer_code_recuperation(code, identifiant):
    """Envoi via l'API HTTP de Resend plutôt qu'en SMTP direct : de nombreux hébergeurs
    (dont Railway) bloquent les ports SMTP sortants (587, 465) par défaut pour lutter
    contre le spam — une requête HTTPS classique passe toujours."""
    api_key = parametre_email("RESEND_API_KEY")
    expediteur = parametre_email("RESEND_FROM") or "onboarding@resend.dev"
    admin_recovery_email = parametre_email("ADMIN_RECOVERY_EMAIL")

    if not api_key or not admin_recovery_email:
        raise RuntimeError(
            "L'envoi d'email n'est pas configuré. Remplissez email_config.py (RESEND_API_KEY, "
            "ADMIN_RECOVERY_EMAIL) puis relancez le serveur."
        )

    corps = (
        f"Code de récupération pour le compte « {identifiant} » sur H.S.E Motors : {code}\n\n"
        "Ce code est valable 15 minutes. Si vous n'êtes pas à l'origine de cette demande, ignorez cet email."
    )
    payload = json.dumps(
        {
            "from": expediteur,
            "to": [admin_recovery_email],
            "subject": f"H.S.E Motors — Code de récupération ({identifiant})",
            "text": corps,
        }
    ).encode("utf-8")

    requete = urllib.request.Request(
        "https://api.resend.com/emails",
        data=payload,
        method="POST",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(requete, timeout=10) as reponse:
            reponse.read()
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Échec de l'envoi de l'email (Resend {e.code}) : {detail}") from e
    except urllib.error.URLError as e:
        raise RuntimeError(f"Échec de l'envoi de l'email : {e.reason}") from e


@app.post("/api/mot-de-passe-oublie")
def demander_code_recuperation():
    data = request.get_json(silent=True) or {}
    identifiant = (data.get("identifiant") or "").strip()
    db = get_db()

    if trop_de_tentatives(db, "mot_de_passe_oublie", identifiant, seuil_identifiant=3, seuil_ip=10):
        return reponse_trop_de_tentatives()

    user = db.execute(
        "SELECT * FROM utilisateurs WHERE identifiant = ? AND role IN ('admin', 'admin_secondaire')", (identifiant,)
    ).fetchone()

    # Réponse identique que le compte existe ou non, pour ne pas révéler les identifiants valides.
    reponse_generique = jsonify(
        ok=True, message="Si ce compte existe, un code de récupération a été envoyé par email."
    )

    enregistrer_tentative(db, "mot_de_passe_oublie", identifiant, reussi=user is not None)
    db.commit()

    if user is None:
        return reponse_generique

    code = f"{secrets.randbelow(1_000_000):06d}"
    expire_le = (datetime.utcnow() + timedelta(minutes=15)).isoformat()
    db.execute(
        "INSERT INTO reinitialisations_mdp (utilisateur_id, code_hash, expire_le) VALUES (?, ?, ?)",
        (user["id"], generate_password_hash(code, method="pbkdf2:sha256"), expire_le),
    )
    db.commit()

    try:
        envoyer_code_recuperation(code, identifiant)
    except Exception as e:
        return jsonify(error=str(e)), 500

    return reponse_generique


@app.post("/api/reinitialiser-mot-de-passe")
def reinitialiser_mot_de_passe():
    data = request.get_json(silent=True) or {}
    identifiant = (data.get("identifiant") or "").strip()
    code = (data.get("code") or "").strip()
    nouveau_mot_de_passe = (data.get("nouveau_mot_de_passe") or "").strip()

    if not nouveau_mot_de_passe:
        return jsonify(error="Le nouveau mot de passe est requis"), 400

    db = get_db()

    if trop_de_tentatives(db, "reinitialiser_mdp", identifiant, seuil_identifiant=5, seuil_ip=15):
        return reponse_trop_de_tentatives()

    user = db.execute(
        "SELECT * FROM utilisateurs WHERE identifiant = ? AND role IN ('admin', 'admin_secondaire')", (identifiant,)
    ).fetchone()
    if user is None:
        enregistrer_tentative(db, "reinitialiser_mdp", identifiant, reussi=False)
        db.commit()
        return jsonify(error="Code invalide ou expiré"), 400

    maintenant = datetime.utcnow().isoformat()
    demandes = db.execute(
        """SELECT * FROM reinitialisations_mdp
           WHERE utilisateur_id = ? AND utilise = 0 AND expire_le > ?
           ORDER BY id DESC""",
        (user["id"], maintenant),
    ).fetchall()

    demande_valide = next((d for d in demandes if check_password_hash(d["code_hash"], code)), None)
    if demande_valide is None:
        enregistrer_tentative(db, "reinitialiser_mdp", identifiant, reussi=False)
        db.commit()
        return jsonify(error="Code invalide ou expiré"), 400

    enregistrer_tentative(db, "reinitialiser_mdp", identifiant, reussi=True)

    db.execute(
        "UPDATE utilisateurs SET password_hash = ? WHERE id = ?",
        (generate_password_hash(nouveau_mot_de_passe, method="pbkdf2:sha256"), user["id"]),
    )
    db.execute("UPDATE reinitialisations_mdp SET utilise = 1 WHERE id = ?", (demande_valide["id"],))
    db.commit()
    return jsonify(ok=True)


@app.get("/api/me")
@login_required
def me():
    return jsonify(user_public(g.user, get_db()))


@app.put("/api/mon-mot-de-passe")
@login_required
def changer_mon_mot_de_passe():
    """Changement de mot de passe en libre-service, réservé aux deux comptes admin
    (le mot de passe des commerciaux et animateurs reste géré par l'admin)."""
    if not est_type_admin(g.user):
        return jsonify(error="Réservé aux administrateurs"), 403

    data = request.get_json(silent=True) or {}
    mot_de_passe_actuel = data.get("mot_de_passe_actuel") or ""
    nouveau_mot_de_passe = (data.get("nouveau_mot_de_passe") or "").strip()

    if not check_password_hash(g.user["password_hash"], mot_de_passe_actuel):
        return jsonify(error="Mot de passe actuel incorrect"), 400
    if len(nouveau_mot_de_passe) < 6:
        return jsonify(error="Le nouveau mot de passe doit faire au moins 6 caractères"), 400

    db = get_db()
    db.execute(
        "UPDATE utilisateurs SET password_hash = ? WHERE id = ?",
        (generate_password_hash(nouveau_mot_de_passe, method="pbkdf2:sha256"), g.user["id"]),
    )
    db.commit()
    return jsonify(ok=True)


@app.put("/api/preferences/kpi-order")
@login_required
def sauvegarder_ordre_kpi():
    data = request.get_json(silent=True) or {}
    ordre = data.get("order")
    if not isinstance(ordre, list) or not all(isinstance(x, str) for x in ordre):
        return jsonify(error="Format d'ordre invalide"), 400

    db = get_db()
    db.execute("UPDATE utilisateurs SET kpi_order = ? WHERE id = ?", (json.dumps(ordre), g.user["id"]))
    db.commit()
    return jsonify(ok=True, order=ordre)


@app.put("/api/preferences/kpi-reduits")
@login_required
def sauvegarder_kpi_reduits():
    data = request.get_json(silent=True) or {}
    reduits = data.get("reduits")
    if not isinstance(reduits, list) or not all(isinstance(x, str) for x in reduits):
        return jsonify(error="Format invalide"), 400

    db = get_db()
    db.execute("UPDATE utilisateurs SET kpi_reduits = ? WHERE id = ?", (json.dumps(reduits), g.user["id"]))
    db.commit()
    return jsonify(ok=True, reduits=reduits)


# ---------- Routes: entreprises ----------

def entreprise_dict(row, db):
    mois_courant = datetime.utcnow().strftime("%Y-%m")
    ca_ht_brut_mois = ca_ht_brut_mensuel_entreprise(db, row["id"], mois_courant)
    objectif = row["objectif_ca_ht_brut"]
    pourcentage_atteinte = (ca_ht_brut_mois / objectif * 100) if objectif else None
    return {
        "id": row["id"],
        "nom": row["nom"],
        "couleur": row["couleur"],
        "objectif_ca_ht_brut": objectif,
        "ca_ht_brut_mois_courant": ca_ht_brut_mois,
        "pourcentage_atteinte": pourcentage_atteinte,
    }


@app.get("/api/entreprises")
@login_required
def list_entreprises():
    db = get_db()
    rows = db.execute("SELECT * FROM entreprises ORDER BY nom").fetchall()
    return jsonify([entreprise_dict(r, db) for r in rows])


@app.post("/api/entreprises")
@admin_required
def create_entreprise():
    data = request.get_json(silent=True) or {}
    nom = (data.get("nom") or "").strip()
    couleur = (data.get("couleur") or "#8B93A1").strip()
    if not nom:
        return jsonify(error="Le nom est requis"), 400
    objectif, erreur = parser_objectif_optionnel(data)
    if erreur:
        return jsonify(error=erreur), 400
    db = get_db()
    count = db.execute("SELECT COUNT(*) FROM entreprises").fetchone()[0]
    if count >= 5:
        return jsonify(error="Limite de 5 entreprises atteinte"), 400
    cur = db.execute(
        "INSERT INTO entreprises (nom, couleur, objectif_ca_ht_brut) VALUES (?, ?, ?)", (nom, couleur, objectif)
    )
    db.commit()
    row = db.execute("SELECT * FROM entreprises WHERE id = ?", (cur.lastrowid,)).fetchone()
    return jsonify(entreprise_dict(row, db)), 201


@app.put("/api/entreprises/<int:entreprise_id>")
@admin_required
def update_entreprise(entreprise_id):
    db = get_db()
    row = db.execute("SELECT * FROM entreprises WHERE id = ?", (entreprise_id,)).fetchone()
    if row is None:
        return jsonify(error="Entreprise introuvable"), 404

    data = request.get_json(silent=True) or {}
    nom = (data.get("nom") or row["nom"]).strip()
    couleur = (data.get("couleur") or row["couleur"]).strip()
    if not nom:
        return jsonify(error="Le nom est requis"), 400

    if "objectif_ca_ht_brut" in data:
        objectif, erreur = parser_objectif_optionnel(data)
        if erreur:
            return jsonify(error=erreur), 400
    else:
        objectif = row["objectif_ca_ht_brut"]

    db.execute(
        "UPDATE entreprises SET nom = ?, couleur = ?, objectif_ca_ht_brut = ? WHERE id = ?",
        (nom, couleur, objectif, entreprise_id),
    )
    db.commit()
    row = db.execute("SELECT * FROM entreprises WHERE id = ?", (entreprise_id,)).fetchone()
    return jsonify(entreprise_dict(row, db))


# ---------- Routes: commerciaux ----------

SEUIL_TAUX_AUTO = 8000.0
TAUX_AUTO_BAS = 0.20
TAUX_AUTO_HAUT = 0.35

# Bonus avis Google : +5% de commission si le commercial obtient au moins un avis
# par voiture vendue dans le mois. Ne s'applique qu'au taux automatique (20/35%) —
# un taux personnalisé à la main n'en bénéficie pas, l'admin l'ajuste lui-même si besoin.
BONUS_AVIS_GOOGLE = 0.05

# Rétrocommission (ANNEXE 8) : quand un commercial n'a réalisé QUE l'entrée ou QUE la
# sortie du mandat, il touche un taux fixe dédié — basé sur sa propre tranche de CA HT
# net du mois — au lieu du taux 20/35% habituel. Un même commercial qui fait les deux
# reste sur le taux classique : 7,5+12,5=20% et 12,5+22,5=35% retombent d'ailleurs
# exactement sur TAUX_AUTO_BAS/TAUX_AUTO_HAUT.
TAUX_ENTRANT_BAS = 0.075
TAUX_ENTRANT_HAUT = 0.125
TAUX_SORTANT_BAS = 0.125
TAUX_SORTANT_HAUT = 0.225


def total_ht_ligne(row):
    tca = max(0, row["garantie_prix_vendu"] - row["garantie_achat"]) * 18 / 118
    return (
        row["mandat_total"] - row["garantie_achat"] - tca - CASH_SENTINEL - row["nettoyage"] - row["achat_livraison"]
    ) / 1.2


def dossiers_retrocommission_mensuel(db, commercial_id, mois):
    """Dossiers d'un AUTRE commercial où celui-ci est crédité via une rétrocommission
    (entrant ou sortant seul) ce mois-ci. Compte pour sa valeur pleine, comme un
    dossier propre — voir ANNEXE 8."""
    return db.execute(
        """SELECT d.garantie_achat, d.garantie_prix_vendu, d.mandat_total, d.nettoyage, d.achat_livraison
           FROM retrocommissions r JOIN dossiers d ON d.id = r.dossier_id
           WHERE r.commercial_id = ? AND substr(d.date, 1, 7) = ?""",
        (commercial_id, mois),
    ).fetchall()


def ca_ht_net_mensuel(db, commercial_id, mois):
    rows = db.execute(
        "SELECT garantie_achat, garantie_prix_vendu, mandat_total, nettoyage, achat_livraison FROM dossiers "
        "WHERE commercial_id = ? AND substr(date, 1, 7) = ?",
        (commercial_id, mois),
    ).fetchall()
    total = sum(total_ht_ligne(r) for r in rows)
    total += sum(total_ht_ligne(r) for r in dossiers_retrocommission_mensuel(db, commercial_id, mois))
    return total


def ca_ht_brut_mensuel(db, commercial_id, mois):
    total = db.execute(
        "SELECT COALESCE(SUM(mandat_total), 0) FROM dossiers WHERE commercial_id = ? AND substr(date, 1, 7) = ?",
        (commercial_id, mois),
    ).fetchone()[0]
    return total / 1.2


def ca_ht_brut_mensuel_entreprise(db, entreprise_id, mois):
    total = db.execute(
        """SELECT COALESCE(SUM(d.mandat_total), 0) FROM dossiers d
           JOIN commerciaux c ON c.id = d.commercial_id
           WHERE c.entreprise_id = ? AND substr(d.date, 1, 7) = ?""",
        (entreprise_id, mois),
    ).fetchone()[0]
    return total / 1.2


def nb_voitures_vendues_mensuel(db, commercial_id, mois):
    propres = db.execute(
        "SELECT COUNT(*) FROM dossiers WHERE commercial_id = ? AND substr(date, 1, 7) = ?",
        (commercial_id, mois),
    ).fetchone()[0]
    return propres + len(dossiers_retrocommission_mensuel(db, commercial_id, mois))


def avis_obtenus_mensuel(db, commercial_id, mois):
    """Nombre d'avis Google enregistrés pour ce commercial ce mois-ci. Seuls les mois
    déjà clos AVANT la bascule (mois strictement antérieur au mois courant) lisent
    l'ancien compteur (avis_google_legacy) — ils ont déjà servi au calcul de commissions
    passées, donc leur valeur reste figée. Le mois en cours au moment de la bascule
    (même s'il avait une valeur dans l'ancien compteur) et tous les mois suivants
    comptent exclusivement les fiches individuelles de la nouvelle table : sinon un avis
    déjà comptabilisé dans l'ancien système resterait bloqué pour toujours et masquerait
    les nouvelles fiches ajoutées ce mois-là (voir migrer_db)."""
    if mois < datetime.utcnow().strftime("%Y-%m"):
        legacy = db.execute(
            "SELECT avis_obtenus FROM avis_google_legacy WHERE commercial_id = ? AND mois = ?",
            (commercial_id, mois),
        ).fetchone()
        if legacy is not None:
            return legacy["avis_obtenus"]
    row = db.execute(
        """SELECT COUNT(*) AS n FROM avis_google a JOIN dossiers d ON d.id = a.dossier_id
           WHERE a.commercial_id = ? AND substr(d.date, 1, 7) = ?""",
        (commercial_id, mois),
    ).fetchone()
    return row["n"] if row else 0


def bonus_avis_actif(db, commercial_id, mois):
    """Objectif = 1 avis par voiture vendue dans le mois."""
    objectif = nb_voitures_vendues_mensuel(db, commercial_id, mois)
    if objectif <= 0:
        return False
    return avis_obtenus_mensuel(db, commercial_id, mois) >= objectif


def taux_automatique(db, commercial_id, mois):
    ca = ca_ht_net_mensuel(db, commercial_id, mois)
    base = TAUX_AUTO_HAUT if ca > SEUIL_TAUX_AUTO else TAUX_AUTO_BAS
    if bonus_avis_actif(db, commercial_id, mois):
        base += BONUS_AVIS_GOOGLE
    return base


def taux_effectif(db, com, mois):
    if com["taux"] is not None:
        return com["taux"]
    return taux_automatique(db, com["id"], mois)


def taux_partage(db, commercial_id, mois, role):
    """Taux fixe entrant/sortant (ANNEXE 8), pour un dossier où ce commercial n'a
    réalisé que l'entrée ou que la sortie. Toujours basé sur sa tranche automatique
    (jamais sur un taux personnalisé — il n'y en a pas d'équivalent pour ce barème) et
    profite du même bonus avis Google que le taux classique."""
    ca = ca_ht_net_mensuel(db, commercial_id, mois)
    haut = ca > SEUIL_TAUX_AUTO
    if role == "entrant":
        base = TAUX_ENTRANT_HAUT if haut else TAUX_ENTRANT_BAS
    else:
        base = TAUX_SORTANT_HAUT if haut else TAUX_SORTANT_BAS
    if bonus_avis_actif(db, commercial_id, mois):
        base += BONUS_AVIS_GOOGLE
    return base


def commercial_dict(row, db):
    mois_courant = datetime.utcnow().strftime("%Y-%m")
    ca_ht_brut_mois = ca_ht_brut_mensuel(db, row["id"], mois_courant)
    objectif = row["objectif_ca_ht_brut"]
    pourcentage_atteinte = (ca_ht_brut_mois / objectif * 100) if objectif else None
    objectif_avis = nb_voitures_vendues_mensuel(db, row["id"], mois_courant)
    return {
        "id": row["id"],
        "nom": row["nom"],
        "entreprise_id": row["entreprise_id"],
        "taux": row["taux"],
        "taux_auto": row["taux"] is None,
        "taux_effectif": taux_effectif(db, row, mois_courant),
        "actif": bool(row["actif"]),
        "objectif_ca_ht_brut": objectif,
        "ca_ht_brut_mois_courant": ca_ht_brut_mois,
        "pourcentage_atteinte": pourcentage_atteinte,
        "avis_obtenus_mois_courant": avis_obtenus_mensuel(db, row["id"], mois_courant),
        "objectif_avis_mois_courant": objectif_avis,
        "bonus_avis_actif": bonus_avis_actif(db, row["id"], mois_courant),
    }


@app.get("/api/commerciaux")
@login_required
def list_commerciaux():
    db = get_db()
    if peut_tout_voir(g.user):
        rows = db.execute("SELECT * FROM commerciaux WHERE supprime = 0 ORDER BY nom").fetchall()
    else:
        rows = db.execute("SELECT * FROM commerciaux WHERE id = ?", (g.user["commercial_id"],)).fetchall()
    return jsonify([commercial_dict(r, db) for r in rows])


def parser_taux_optionnel(data):
    """Retourne (taux, erreur). taux est None si absent/vide (mode automatique)."""
    if "taux" not in data or data["taux"] in (None, ""):
        return None, None
    try:
        return float(data["taux"]), None
    except (TypeError, ValueError):
        return None, "Taux invalide"


def parser_objectif_optionnel(data):
    """Retourne (objectif, erreur). objectif est None si absent/vide (pas d'objectif défini)."""
    if "objectif_ca_ht_brut" not in data or data["objectif_ca_ht_brut"] in (None, ""):
        return None, None
    try:
        valeur = float(data["objectif_ca_ht_brut"])
    except (TypeError, ValueError):
        return None, "Objectif invalide"
    if valeur <= 0:
        return None, "Objectif invalide"
    return valeur, None


@app.post("/api/commerciaux")
@admin_required
def create_commercial():
    data = request.get_json(silent=True) or {}
    nom = (data.get("nom") or "").strip()
    entreprise_id = data.get("entreprise_id")
    identifiant = (data.get("identifiant") or "").strip()
    mot_de_passe = data.get("mot_de_passe") or ""

    if not nom or not entreprise_id or not identifiant or not mot_de_passe:
        return jsonify(error="Nom, entreprise, identifiant et mot de passe sont requis"), 400

    taux, erreur = parser_taux_optionnel(data)
    if erreur:
        return jsonify(error=erreur), 400
    objectif, erreur = parser_objectif_optionnel(data)
    if erreur:
        return jsonify(error=erreur), 400

    db = get_db()
    if db.execute("SELECT 1 FROM utilisateurs WHERE identifiant = ?", (identifiant,)).fetchone():
        return jsonify(error="Cet identifiant existe déjà"), 400
    if not db.execute("SELECT 1 FROM entreprises WHERE id = ?", (entreprise_id,)).fetchone():
        return jsonify(error="Entreprise inconnue"), 400

    cur = db.execute(
        "INSERT INTO commerciaux (nom, entreprise_id, taux, objectif_ca_ht_brut) VALUES (?, ?, ?, ?)",
        (nom, entreprise_id, taux, objectif),
    )
    commercial_id = cur.lastrowid
    db.execute(
        "INSERT INTO utilisateurs (identifiant, password_hash, role, commercial_id) VALUES (?, ?, 'commercial', ?)",
        (identifiant, generate_password_hash(mot_de_passe, method="pbkdf2:sha256"), commercial_id),
    )
    db.commit()
    row = db.execute("SELECT * FROM commerciaux WHERE id = ?", (commercial_id,)).fetchone()
    return jsonify(commercial_dict(row, db)), 201


@app.put("/api/commerciaux/<int:commercial_id>")
@admin_required
def update_commercial(commercial_id):
    data = request.get_json(silent=True) or {}
    db = get_db()
    row = db.execute("SELECT * FROM commerciaux WHERE id = ?", (commercial_id,)).fetchone()
    if row is None:
        return jsonify(error="Commercial introuvable"), 404

    nom = data.get("nom", row["nom"])
    actif = data.get("actif", bool(row["actif"]))
    entreprise_id = data.get("entreprise_id", row["entreprise_id"])

    if "taux" in data:
        taux, erreur = parser_taux_optionnel(data)
        if erreur:
            return jsonify(error=erreur), 400
    else:
        taux = row["taux"]

    if "objectif_ca_ht_brut" in data:
        objectif, erreur = parser_objectif_optionnel(data)
        if erreur:
            return jsonify(error=erreur), 400
    else:
        objectif = row["objectif_ca_ht_brut"]

    if not db.execute("SELECT 1 FROM entreprises WHERE id = ?", (entreprise_id,)).fetchone():
        return jsonify(error="Entreprise inconnue"), 400

    nouveau_mot_de_passe = (data.get("mot_de_passe") or "").strip()

    db.execute(
        "UPDATE commerciaux SET nom = ?, taux = ?, actif = ?, entreprise_id = ?, objectif_ca_ht_brut = ? WHERE id = ?",
        (nom, taux, 1 if actif else 0, entreprise_id, objectif, commercial_id),
    )
    if nouveau_mot_de_passe:
        db.execute(
            "UPDATE utilisateurs SET password_hash = ? WHERE commercial_id = ?",
            (generate_password_hash(nouveau_mot_de_passe, method="pbkdf2:sha256"), commercial_id),
        )
    db.commit()
    row = db.execute("SELECT * FROM commerciaux WHERE id = ?", (commercial_id,)).fetchone()
    return jsonify(commercial_dict(row, db))


@app.delete("/api/commerciaux/<int:commercial_id>")
@admin_required
def delete_commercial(commercial_id):
    """Supprime la fiche du commercial (il disparaît de la liste, ne peut plus se
    connecter, ne peut plus se voir assigner de nouveau dossier) sans toucher à ses
    dossiers existants : leur historique (CA, commissions, export Excel) reste intact
    puisque la ligne commerciaux elle-même n'est pas effacée, seulement masquée."""
    db = get_db()
    if not db.execute("SELECT 1 FROM commerciaux WHERE id = ?", (commercial_id,)).fetchone():
        return jsonify(error="Commercial introuvable"), 404
    db.execute("UPDATE commerciaux SET supprime = 1, actif = 0 WHERE id = ?", (commercial_id,))
    db.commit()
    return jsonify(ok=True)


# ---------- Routes: animateurs (accès lecture seule, toutes entreprises) ----------

def animateur_dict(row):
    return {"id": row["id"], "identifiant": row["identifiant"], "nom": row["nom"]}


@app.get("/api/animateurs")
@admin_required
def list_animateurs():
    db = get_db()
    rows = db.execute("SELECT * FROM utilisateurs WHERE role = 'animateur' ORDER BY nom").fetchall()
    return jsonify([animateur_dict(r) for r in rows])


@app.post("/api/animateurs")
@admin_required
def create_animateur():
    data = request.get_json(silent=True) or {}
    nom = (data.get("nom") or "").strip()
    identifiant = (data.get("identifiant") or "").strip()
    mot_de_passe = data.get("mot_de_passe") or ""

    if not nom or not identifiant or not mot_de_passe:
        return jsonify(error="Nom, identifiant et mot de passe sont requis"), 400

    db = get_db()
    if db.execute("SELECT 1 FROM utilisateurs WHERE identifiant = ?", (identifiant,)).fetchone():
        return jsonify(error="Cet identifiant existe déjà"), 400

    cur = db.execute(
        "INSERT INTO utilisateurs (identifiant, password_hash, role, nom) VALUES (?, ?, 'animateur', ?)",
        (identifiant, generate_password_hash(mot_de_passe, method="pbkdf2:sha256"), nom),
    )
    db.commit()
    row = db.execute("SELECT * FROM utilisateurs WHERE id = ?", (cur.lastrowid,)).fetchone()
    return jsonify(animateur_dict(row)), 201


# ---------- Routes: dossiers ----------

CASH_SENTINEL = 43.20


def dossier_dict(row, com, ent, db):
    tca = max(0, row["garantie_prix_vendu"] - row["garantie_achat"]) * 18 / 118
    total_ht = (
        row["mandat_total"] - row["garantie_achat"] - tca - CASH_SENTINEL - row["nettoyage"] - row["achat_livraison"]
    ) / 1.2
    mois = row["date"][:7]
    role_commercial = row["role_commercial"]

    if role_commercial == "les_deux":
        taux = taux_effectif(db, com, mois)
    else:
        taux = taux_partage(db, com["id"], mois, role_commercial)
    commission = total_ht * taux

    retro_row = db.execute("SELECT * FROM retrocommissions WHERE dossier_id = ?", (row["id"],)).fetchone()
    retrocommission = None
    if retro_row is not None:
        retro_com = db.execute("SELECT * FROM commerciaux WHERE id = ?", (retro_row["commercial_id"],)).fetchone()
        retro_taux = taux_partage(db, retro_row["commercial_id"], mois, retro_row["role"])
        retrocommission = {
            "commercial_id": retro_row["commercial_id"],
            "commercial_nom": retro_com["nom"] if retro_com else "—",
            "role": retro_row["role"],
            "commission": total_ht * retro_taux,
        }

    return {
        "id": row["id"],
        "commercial_id": row["commercial_id"],
        "commercial_nom": com["nom"],
        "entreprise_id": ent["id"],
        "entreprise_nom": ent["nom"],
        "entreprise_couleur": ent["couleur"],
        "date": row["date"],
        "client": row["client"],
        "voiture": row["voiture"],
        "plaque": row["plaque"],
        "garantie_achat": row["garantie_achat"],
        "garantie_prix_vendu": row["garantie_prix_vendu"],
        "mandat_total": row["mandat_total"],
        "frais_intermediation": row["frais_intermediation"],
        "nettoyage": row["nettoyage"],
        "commission_agence": row["commission_agence"],
        "achat_livraison": row["achat_livraison"],
        "vente_livraison": row["vente_livraison"],
        "tca": tca,
        "cash_sentinel": CASH_SENTINEL,
        "total_ht": total_ht,
        "role_commercial": role_commercial,
        "retrocommission": retrocommission,
        "commission": commission,
    }


def enrich(db, rows):
    result = []
    for row in rows:
        com = db.execute("SELECT * FROM commerciaux WHERE id = ?", (row["commercial_id"],)).fetchone()
        ent = db.execute("SELECT * FROM entreprises WHERE id = ?", (com["entreprise_id"],)).fetchone()
        result.append(dossier_dict(row, com, ent, db))
    return result


@app.get("/api/dossiers")
@login_required
def list_dossiers():
    db = get_db()
    if peut_tout_voir(g.user):
        entreprise_id = request.args.get("entreprise_id")
        if entreprise_id and entreprise_id != "toutes":
            rows = db.execute(
                """SELECT d.* FROM dossiers d
                   JOIN commerciaux c ON c.id = d.commercial_id
                   WHERE c.entreprise_id = ? ORDER BY d.date DESC""",
                (entreprise_id,),
            ).fetchall()
        else:
            rows = db.execute("SELECT * FROM dossiers ORDER BY date DESC").fetchall()
    else:
        rows = db.execute(
            "SELECT * FROM dossiers WHERE commercial_id = ? ORDER BY date DESC",
            (g.user["commercial_id"],),
        ).fetchall()
    return jsonify(enrich(db, rows))


def resolve_commercial_id(data):
    if est_type_admin(g.user):
        commercial_id = data.get("commercial_id")
        if not commercial_id:
            return None, (jsonify(error="commercial_id requis"), 400)
        return commercial_id, None
    return g.user["commercial_id"], None


def mandat_total_calcule(frais_intermediation, commission_agence, garantie_prix_vendu, vente_livraison):
    """Le mandat total n'est plus saisi : c'est la somme de ces montants."""
    return frais_intermediation + commission_agence + garantie_prix_vendu + vente_livraison


def parser_role_et_retrocommission(data, commercial_id, db):
    """Valide le rôle du commercial du dossier et, si le mandat est partagé, la
    rétrocommission (l'autre commercial + son rôle, forcément complémentaire).
    Retourne (role_commercial, retro_commercial_id, retro_role, erreur) — erreur est
    None si tout est valide, sinon un tuple (réponse jsonify, code) à retourner tel quel."""
    role_commercial = data.get("role_commercial", "les_deux")
    if role_commercial not in ("les_deux", "entrant", "sortant"):
        return None, None, None, (jsonify(error="Rôle invalide"), 400)

    if role_commercial == "les_deux":
        return role_commercial, None, None, None

    retro = data.get("retrocommission") or {}
    retro_commercial_id = retro.get("commercial_id")
    retro_role = retro.get("role")
    role_oppose = "sortant" if role_commercial == "entrant" else "entrant"

    if not retro_commercial_id:
        return None, None, None, (jsonify(error="Rétrocommission : commercial requis"), 400)
    if retro_role != role_oppose:
        return None, None, None, (jsonify(error="Rôle de rétrocommission incohérent"), 400)
    if int(retro_commercial_id) == int(commercial_id):
        return None, None, None, (jsonify(error="La rétrocommission doit concerner un autre commercial"), 400)
    if not db.execute("SELECT 1 FROM commerciaux WHERE id = ?", (retro_commercial_id,)).fetchone():
        return None, None, None, (jsonify(error="Commercial de rétrocommission inconnu"), 400)

    return role_commercial, retro_commercial_id, retro_role, None


def sauvegarder_retrocommission(db, dossier_id, retro_commercial_id, retro_role):
    """Remplace la rétrocommission éventuelle d'un dossier (au plus une par dossier)."""
    db.execute("DELETE FROM retrocommissions WHERE dossier_id = ?", (dossier_id,))
    if retro_commercial_id is not None:
        db.execute(
            "INSERT INTO retrocommissions (dossier_id, commercial_id, role) VALUES (?, ?, ?)",
            (dossier_id, retro_commercial_id, retro_role),
        )


@app.post("/api/dossiers")
@ecriture_requise
def create_dossier():
    data = request.get_json(silent=True) or {}
    commercial_id, err = resolve_commercial_id(data)
    if err:
        return err

    client = (data.get("client") or "").strip()
    voiture = (data.get("voiture") or "").strip()
    frais_intermediation = float(data.get("frais_intermediation") or 0)
    commission_agence = float(data.get("commission_agence") or 0)
    garantie_prix_vendu = float(data.get("garantie_prix_vendu") or 0)
    vente_livraison = float(data.get("vente_livraison") or 0)
    mandat_total = mandat_total_calcule(frais_intermediation, commission_agence, garantie_prix_vendu, vente_livraison)
    if not client or not voiture or mandat_total <= 0:
        return jsonify(error="Client, voiture et mandat total (>0) sont requis"), 400

    db = get_db()
    if not db.execute("SELECT 1 FROM commerciaux WHERE id = ?", (commercial_id,)).fetchone():
        return jsonify(error="Commercial inconnu"), 400

    role_commercial, retro_commercial_id, retro_role, erreur = parser_role_et_retrocommission(data, commercial_id, db)
    if erreur:
        return erreur

    cur = db.execute(
        """INSERT INTO dossiers
           (commercial_id, date, client, voiture, plaque, garantie_achat, garantie_prix_vendu, mandat_total, frais_intermediation, nettoyage, commission_agence, achat_livraison, vente_livraison, role_commercial)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            commercial_id,
            data.get("date") or str(date.today()),
            client,
            voiture,
            (data.get("plaque") or "").strip().upper(),
            float(data.get("garantie_achat") or 0),
            garantie_prix_vendu,
            mandat_total,
            frais_intermediation,
            float(data.get("nettoyage") or 0),
            commission_agence,
            float(data.get("achat_livraison") or 0),
            vente_livraison,
            role_commercial,
        ),
    )
    sauvegarder_retrocommission(db, cur.lastrowid, retro_commercial_id, retro_role)
    db.commit()
    row = db.execute("SELECT * FROM dossiers WHERE id = ?", (cur.lastrowid,)).fetchone()
    return jsonify(enrich(db, [row])[0]), 201


def load_dossier_for_user(db, dossier_id):
    row = db.execute("SELECT * FROM dossiers WHERE id = ?", (dossier_id,)).fetchone()
    if row is None:
        return None
    if not est_type_admin(g.user) and row["commercial_id"] != g.user["commercial_id"]:
        return None
    return row


@app.put("/api/dossiers/<int:dossier_id>")
@ecriture_requise
def update_dossier(dossier_id):
    db = get_db()
    row = load_dossier_for_user(db, dossier_id)
    if row is None:
        return jsonify(error="Dossier introuvable"), 404

    data = request.get_json(silent=True) or {}
    commercial_id = row["commercial_id"]
    if est_type_admin(g.user) and data.get("commercial_id"):
        commercial_id = data["commercial_id"]

    frais_intermediation = float(data.get("frais_intermediation", row["frais_intermediation"]) or 0)
    commission_agence = float(data.get("commission_agence", row["commission_agence"]) or 0)
    garantie_prix_vendu = float(data.get("garantie_prix_vendu", row["garantie_prix_vendu"]) or 0)
    vente_livraison = float(data.get("vente_livraison", row["vente_livraison"]) or 0)
    mandat_total = mandat_total_calcule(frais_intermediation, commission_agence, garantie_prix_vendu, vente_livraison)

    data.setdefault("role_commercial", row["role_commercial"])
    role_commercial, retro_commercial_id, retro_role, erreur = parser_role_et_retrocommission(data, commercial_id, db)
    if erreur:
        return erreur

    db.execute(
        """UPDATE dossiers SET commercial_id = ?, date = ?, client = ?, voiture = ?, plaque = ?,
           garantie_achat = ?, garantie_prix_vendu = ?, mandat_total = ?, frais_intermediation = ?, nettoyage = ?,
           commission_agence = ?, achat_livraison = ?, vente_livraison = ?, role_commercial = ? WHERE id = ?""",
        (
            commercial_id,
            data.get("date", row["date"]),
            (data.get("client", row["client"]) or "").strip(),
            (data.get("voiture", row["voiture"]) or "").strip(),
            (data.get("plaque", row["plaque"]) or "").strip().upper(),
            float(data.get("garantie_achat", row["garantie_achat"]) or 0),
            garantie_prix_vendu,
            mandat_total,
            frais_intermediation,
            float(data.get("nettoyage", row["nettoyage"]) or 0),
            commission_agence,
            float(data.get("achat_livraison", row["achat_livraison"]) or 0),
            vente_livraison,
            role_commercial,
            dossier_id,
        ),
    )
    sauvegarder_retrocommission(db, dossier_id, retro_commercial_id, retro_role)
    db.commit()
    row = db.execute("SELECT * FROM dossiers WHERE id = ?", (dossier_id,)).fetchone()
    return jsonify(enrich(db, [row])[0])


@app.delete("/api/dossiers/<int:dossier_id>")
@ecriture_requise
def delete_dossier(dossier_id):
    db = get_db()
    row = load_dossier_for_user(db, dossier_id)
    if row is None:
        return jsonify(error="Dossier introuvable"), 404
    db.execute("DELETE FROM retrocommissions WHERE dossier_id = ?", (dossier_id,))
    db.execute("DELETE FROM dossiers WHERE id = ?", (dossier_id,))
    db.commit()
    return jsonify(ok=True)


@app.get("/api/retrocommissions")
@login_required
def list_retrocommissions():
    """Les lignes de rétrocommission créditées à l'utilisateur courant (ou à tout le
    monde pour l'admin/animateur) : le dossier appartient à un AUTRE commercial, mais
    celui-ci touche sa part (entrant ou sortant) — voir ANNEXE 8."""
    db = get_db()
    requete = """SELECT r.id, r.dossier_id, r.commercial_id, r.role,
                        d.date, d.client, d.voiture, d.plaque,
                        d.garantie_achat, d.garantie_prix_vendu, d.mandat_total, d.nettoyage, d.achat_livraison
                 FROM retrocommissions r JOIN dossiers d ON d.id = r.dossier_id"""
    if peut_tout_voir(g.user):
        rows = db.execute(requete + " ORDER BY d.date DESC").fetchall()
    else:
        rows = db.execute(
            requete + " WHERE r.commercial_id = ? ORDER BY d.date DESC", (g.user["commercial_id"],)
        ).fetchall()

    result = []
    for row in rows:
        com = db.execute("SELECT * FROM commerciaux WHERE id = ?", (row["commercial_id"],)).fetchone()
        ent = db.execute("SELECT * FROM entreprises WHERE id = ?", (com["entreprise_id"],)).fetchone() if com else None
        taux = taux_partage(db, row["commercial_id"], row["date"][:7], row["role"])
        result.append(
            {
                "id": row["id"],
                "dossier_id": row["dossier_id"],
                "date": row["date"],
                "client": row["client"],
                "voiture": row["voiture"],
                "plaque": row["plaque"],
                "commercial_id": row["commercial_id"],
                "commercial_nom": com["nom"] if com else "—",
                "entreprise_id": ent["id"] if ent else None,
                "entreprise_nom": ent["nom"] if ent else None,
                "entreprise_couleur": ent["couleur"] if ent else None,
                "role": row["role"],
                "commission": total_ht_ligne(row) * taux,
            }
        )
    return jsonify(result)


# ---------- Routes: avis Google ----------
#
# Chaque avis est une fiche individuelle (nom de la personne, vendeur/acheteur — un
# champ purement informatif, sans impact sur le calcul — rattachée à un dossier), pas
# un simple compteur : le commercial concerné ou l'admin peut en ajouter un. Le mois
# pris en compte pour le bonus est celui du dossier lié, jamais saisi à part.

def dossiers_lies_au_commercial(db, commercial_id):
    """Dossiers sur lesquels un avis peut être rattaché pour ce commercial : les siens
    en propre, ou ceux d'un autre commercial où il est crédité via une rétrocommission."""
    return db.execute(
        """SELECT id FROM dossiers WHERE commercial_id = ?
           UNION
           SELECT dossier_id FROM retrocommissions WHERE commercial_id = ?""",
        (commercial_id, commercial_id),
    ).fetchall()


def avis_dict(row):
    return {
        "id": row["id"],
        "dossier_id": row["dossier_id"],
        "commercial_id": row["commercial_id"],
        "commercial_nom": row["commercial_nom"],
        "entreprise_id": row["entreprise_id"],
        "entreprise_nom": row["entreprise_nom"],
        "entreprise_couleur": row["entreprise_couleur"],
        "date": row["date"],
        "client": row["client"],
        "voiture": row["voiture"],
        "nom_client": row["nom_client"],
        "role": row["role"],
        "cree_le": row["cree_le"],
    }


@app.get("/api/avis")
@login_required
def list_avis():
    db = get_db()
    requete = """SELECT a.id, a.dossier_id, a.commercial_id, a.nom_client, a.role, a.cree_le,
                        c.nom AS commercial_nom, e.id AS entreprise_id, e.nom AS entreprise_nom,
                        e.couleur AS entreprise_couleur, d.date, d.client, d.voiture
                 FROM avis_google a
                 JOIN commerciaux c ON c.id = a.commercial_id
                 JOIN entreprises e ON e.id = c.entreprise_id
                 JOIN dossiers d ON d.id = a.dossier_id"""
    if peut_tout_voir(g.user):
        rows = db.execute(requete + " ORDER BY d.date DESC, a.id DESC").fetchall()
    else:
        rows = db.execute(
            requete + " WHERE a.commercial_id = ? ORDER BY d.date DESC, a.id DESC",
            (g.user["commercial_id"],),
        ).fetchall()
    return jsonify([avis_dict(r) for r in rows])


@app.post("/api/avis")
@ecriture_requise
def create_avis():
    data = request.get_json(silent=True) or {}
    db = get_db()

    commercial_id, err = resolve_commercial_id(data)
    if err:
        return err
    if not db.execute("SELECT 1 FROM commerciaux WHERE id = ?", (commercial_id,)).fetchone():
        return jsonify(error="Commercial inconnu"), 400

    dossier_id = data.get("dossier_id")
    nom_client = (data.get("nom_client") or "").strip()
    role = data.get("role")

    if not dossier_id or not nom_client or role not in ("vendeur", "acheteur"):
        return jsonify(error="Dossier, nom et rôle (vendeur/acheteur) sont requis"), 400

    dossiers_valides = {r["id"] for r in dossiers_lies_au_commercial(db, commercial_id)}
    if int(dossier_id) not in dossiers_valides:
        return jsonify(error="Ce dossier n'est pas rattaché à ce commercial"), 400

    try:
        cur = db.execute(
            "INSERT INTO avis_google (dossier_id, commercial_id, nom_client, role, cree_le) VALUES (?, ?, ?, ?, ?)",
            (dossier_id, commercial_id, nom_client, role, datetime.utcnow().isoformat()),
        )
    except sqlite3.IntegrityError:
        return jsonify(error="Un avis existe déjà pour ce rôle sur ce dossier"), 400
    db.commit()

    row = db.execute(
        """SELECT a.id, a.dossier_id, a.commercial_id, a.nom_client, a.role, a.cree_le,
                  c.nom AS commercial_nom, e.id AS entreprise_id, e.nom AS entreprise_nom,
                  e.couleur AS entreprise_couleur, d.date, d.client, d.voiture
           FROM avis_google a
           JOIN commerciaux c ON c.id = a.commercial_id
           JOIN entreprises e ON e.id = c.entreprise_id
           JOIN dossiers d ON d.id = a.dossier_id
           WHERE a.id = ?""",
        (cur.lastrowid,),
    ).fetchone()
    return jsonify(avis_dict(row)), 201


@app.delete("/api/avis/<int:avis_id>")
@ecriture_requise
def delete_avis(avis_id):
    db = get_db()
    row = db.execute("SELECT * FROM avis_google WHERE id = ?", (avis_id,)).fetchone()
    if row is None:
        return jsonify(error="Avis introuvable"), 404
    if not est_type_admin(g.user) and row["commercial_id"] != g.user["commercial_id"]:
        return jsonify(error="Avis introuvable"), 404
    db.execute("DELETE FROM avis_google WHERE id = ?", (avis_id,))
    db.commit()
    return jsonify(ok=True)


# ---------- Routes: refacturation ----------
#
# Suivi mensuel (mois courant uniquement, comme les avis Google) du volume de leads
# par commercial pour Call center et Leads Meta, et du montant Leboncoin. Le nombre
# est saisi par le commercial lui-même ; les montants (prix unitaire des deux
# premières catégories, montant Leboncoin par commercial) sont fixés par l'admin.


def get_tarifs_refacturation(db):
    rows = db.execute("SELECT categorie, prix_unitaire FROM tarifs_refacturation").fetchall()
    return {r["categorie"]: r["prix_unitaire"] for r in rows}


def refacturation_ligne(row, tarifs):
    montant_call_center = row["nombre_call_center"] * tarifs.get("call_center", 0)
    montant_leads_meta = row["nombre_leads_meta"] * tarifs.get("leads_meta", 0)
    return {
        "nombre_call_center": row["nombre_call_center"],
        "nombre_leads_meta": row["nombre_leads_meta"],
        "montant_call_center": montant_call_center,
        "montant_leads_meta": montant_leads_meta,
        "montant_leboncoin": row["montant_leboncoin"],
        "total": montant_call_center + montant_leads_meta + row["montant_leboncoin"],
    }


@app.get("/api/refacturation")
@login_required
def get_refacturation():
    db = get_db()
    mois = datetime.utcnow().strftime("%Y-%m")
    tarifs = get_tarifs_refacturation(db)

    if peut_tout_voir(g.user):
        commerciaux_rows = db.execute("SELECT * FROM commerciaux WHERE supprime = 0 ORDER BY nom").fetchall()
    else:
        commerciaux_rows = db.execute(
            "SELECT * FROM commerciaux WHERE id = ?", (g.user["commercial_id"],)
        ).fetchall()

    vide = {"nombre_call_center": 0, "nombre_leads_meta": 0, "montant_leboncoin": 0}
    lignes = []
    for com in commerciaux_rows:
        row = db.execute(
            "SELECT * FROM refacturations WHERE commercial_id = ? AND mois = ?", (com["id"], mois)
        ).fetchone()
        ligne = refacturation_ligne(row if row is not None else vide, tarifs)
        ligne["commercial_id"] = com["id"]
        ligne["commercial_nom"] = com["nom"]
        lignes.append(ligne)

    return jsonify(mois=mois, tarifs=tarifs, lignes=lignes)


@app.put("/api/refacturation/nombre")
@ecriture_requise
def update_refacturation_nombre():
    data = request.get_json(silent=True) or {}
    categorie = data.get("categorie")
    if categorie not in ("call_center", "leads_meta"):
        return jsonify(error="Catégorie invalide"), 400

    if est_type_admin(g.user):
        commercial_id = data.get("commercial_id")
        if not commercial_id:
            return jsonify(error="commercial_id requis"), 400
    else:
        commercial_id = g.user["commercial_id"]

    try:
        nombre = int(data.get("nombre") or 0)
    except (TypeError, ValueError):
        return jsonify(error="Nombre invalide"), 400
    if nombre < 0:
        return jsonify(error="Nombre invalide"), 400

    db = get_db()
    if not db.execute("SELECT 1 FROM commerciaux WHERE id = ?", (commercial_id,)).fetchone():
        return jsonify(error="Commercial inconnu"), 400

    mois = datetime.utcnow().strftime("%Y-%m")
    colonne = "nombre_call_center" if categorie == "call_center" else "nombre_leads_meta"
    db.execute(
        f"""INSERT INTO refacturations (commercial_id, mois, {colonne}) VALUES (?, ?, ?)
            ON CONFLICT(commercial_id, mois) DO UPDATE SET {colonne} = excluded.{colonne}""",
        (commercial_id, mois, nombre),
    )
    db.commit()
    return jsonify(ok=True)


@app.put("/api/refacturation/leboncoin")
@admin_required
def update_refacturation_leboncoin():
    data = request.get_json(silent=True) or {}
    commercial_id = data.get("commercial_id")
    if not commercial_id:
        return jsonify(error="commercial_id requis"), 400
    try:
        montant = float(data.get("montant") or 0)
    except (TypeError, ValueError):
        return jsonify(error="Montant invalide"), 400

    db = get_db()
    if not db.execute("SELECT 1 FROM commerciaux WHERE id = ?", (commercial_id,)).fetchone():
        return jsonify(error="Commercial inconnu"), 400

    mois = datetime.utcnow().strftime("%Y-%m")
    db.execute(
        """INSERT INTO refacturations (commercial_id, mois, montant_leboncoin) VALUES (?, ?, ?)
           ON CONFLICT(commercial_id, mois) DO UPDATE SET montant_leboncoin = excluded.montant_leboncoin""",
        (commercial_id, mois, montant),
    )
    db.commit()
    return jsonify(ok=True)


@app.put("/api/refacturation/tarifs")
@admin_complet_requis
def update_refacturation_tarif():
    data = request.get_json(silent=True) or {}
    categorie = data.get("categorie")
    if categorie not in ("call_center", "leads_meta"):
        return jsonify(error="Catégorie invalide"), 400
    try:
        prix = float(data.get("prix_unitaire") or 0)
    except (TypeError, ValueError):
        return jsonify(error="Prix invalide"), 400

    db = get_db()
    db.execute(
        "INSERT INTO tarifs_refacturation (categorie, prix_unitaire) VALUES (?, ?) "
        "ON CONFLICT(categorie) DO UPDATE SET prix_unitaire = excluded.prix_unitaire",
        (categorie, prix),
    )
    db.commit()
    return jsonify(ok=True)


# ---------- Routes: charges fixes (admin uniquement) ----------
#
# Liste de postes de charges (nom + montant), propre à chaque entreprise et remise à
# zéro chaque mois : une charge n'existe que pour le mois où elle a été créée, et ne
# peut être créée que dans le mois courant. Les mois passés restent consultables (comme
# le CA) mais ne sont plus modifiables, pour garder un historique fiable.


def mois_courant_str():
    return datetime.utcnow().strftime("%Y-%m")


def charge_fixe_dict(row, ent):
    return {
        "id": row["id"],
        "entreprise_id": row["entreprise_id"],
        "entreprise_nom": ent["nom"] if ent else None,
        "entreprise_couleur": ent["couleur"] if ent else None,
        "mois": row["mois"],
        "nom": row["nom"],
        "montant": row["montant"],
    }


def enrichir_charge_fixe(db, row):
    ent = db.execute("SELECT * FROM entreprises WHERE id = ?", (row["entreprise_id"],)).fetchone()
    return charge_fixe_dict(row, ent)


@app.get("/api/charges-fixes")
@admin_required
def list_charges_fixes():
    db = get_db()
    rows = db.execute("SELECT * FROM charges_fixes ORDER BY mois DESC, id").fetchall()
    return jsonify([enrichir_charge_fixe(db, r) for r in rows])


@app.post("/api/charges-fixes")
@admin_complet_requis
def create_charge_fixe():
    data = request.get_json(silent=True) or {}
    nom = (data.get("nom") or "").strip()
    entreprise_id = data.get("entreprise_id")
    try:
        montant = float(data.get("montant") or 0)
    except (TypeError, ValueError):
        return jsonify(error="Montant invalide"), 400
    if not nom or montant <= 0:
        return jsonify(error="Nom et montant (>0) sont requis"), 400
    if not entreprise_id:
        return jsonify(error="Entreprise requise"), 400

    db = get_db()
    if not db.execute("SELECT 1 FROM entreprises WHERE id = ?", (entreprise_id,)).fetchone():
        return jsonify(error="Entreprise inconnue"), 400

    cur = db.execute(
        "INSERT INTO charges_fixes (entreprise_id, mois, nom, montant) VALUES (?, ?, ?, ?)",
        (entreprise_id, mois_courant_str(), nom, montant),
    )
    db.commit()
    row = db.execute("SELECT * FROM charges_fixes WHERE id = ?", (cur.lastrowid,)).fetchone()
    return jsonify(enrichir_charge_fixe(db, row)), 201


@app.put("/api/charges-fixes/<int:charge_id>")
@admin_complet_requis
def update_charge_fixe(charge_id):
    db = get_db()
    row = db.execute("SELECT * FROM charges_fixes WHERE id = ?", (charge_id,)).fetchone()
    if row is None:
        return jsonify(error="Charge introuvable"), 404
    if row["mois"] != mois_courant_str():
        return jsonify(error="Seules les charges du mois en cours peuvent être modifiées"), 400

    data = request.get_json(silent=True) or {}
    nom = (data.get("nom", row["nom"]) or "").strip()
    try:
        montant = float(data.get("montant", row["montant"]) or 0)
    except (TypeError, ValueError):
        return jsonify(error="Montant invalide"), 400
    if not nom or montant <= 0:
        return jsonify(error="Nom et montant (>0) sont requis"), 400

    db.execute("UPDATE charges_fixes SET nom = ?, montant = ? WHERE id = ?", (nom, montant, charge_id))
    db.commit()
    row = db.execute("SELECT * FROM charges_fixes WHERE id = ?", (charge_id,)).fetchone()
    return jsonify(enrichir_charge_fixe(db, row))


@app.delete("/api/charges-fixes/<int:charge_id>")
@admin_complet_requis
def delete_charge_fixe(charge_id):
    db = get_db()
    row = db.execute("SELECT * FROM charges_fixes WHERE id = ?", (charge_id,)).fetchone()
    if row is None:
        return jsonify(error="Charge introuvable"), 404
    if row["mois"] != mois_courant_str():
        return jsonify(error="Seules les charges du mois en cours peuvent être supprimées"), 400
    db.execute("DELETE FROM charges_fixes WHERE id = ?", (charge_id,))
    db.commit()
    return jsonify(ok=True)


# ---------- Routes: export Excel ----------
#
# Le fichier généré reproduit la structure du tableau de suivi papier/Excel
# existant (une feuille par commercial, colonnes VOITURE/PLAQUE/NOM/FRAIS
# INTER/MOYEN DE PAIEMENT/DATE/MONTANT CASHSENTINEL aux mêmes positions),
# afin qu'un copier-coller depuis ce fichier s'aligne directement dans le
# tableau réel. "MOYEN DE PAIEMENT" n'est pas suivi par l'app : la colonne
# reste vide, à compléter à la main. C'est un NOUVEAU fichier à chaque fois
# (l'app ne touche jamais le fichier de suivi réel de l'utilisateur).

EXPORT_ENTETES = {
    1: "VOITURE",
    3: "PLAQUE",
    4: "NOM",
    5: "FRAIS INTER",
    6: "MOYEN DE PAIEMENT",
    7: "DATE",
    8: "MONTANT CASHSENTINEL",
}


def nom_feuille_unique(nom_souhaite, noms_deja_pris):
    interdits = set('[]:*?/\\')
    base = "".join(c for c in nom_souhaite if c not in interdits).strip()[:31] or "Commercial"
    nom = base
    n = 2
    while nom in noms_deja_pris:
        suffixe = f" ({n})"
        nom = base[: 31 - len(suffixe)] + suffixe
        n += 1
    noms_deja_pris.add(nom)
    return nom


def construire_classeur_export(commerciaux_rows, dossiers_par_commercial):
    wb = Workbook()
    wb.remove(wb.active)
    noms_deja_pris = set()
    for com in commerciaux_rows:
        dossiers = dossiers_par_commercial.get(com["id"]) or []
        if not dossiers:
            continue
        ws = wb.create_sheet(nom_feuille_unique(com["nom"], noms_deja_pris))
        for col, entete in EXPORT_ENTETES.items():
            ws.cell(row=1, column=col, value=entete)
        for i, d in enumerate(dossiers, start=2):
            ws.cell(row=i, column=1, value=d["voiture"])
            ws.cell(row=i, column=3, value=d["plaque"])
            ws.cell(row=i, column=4, value=d["client"])
            ws.cell(row=i, column=5, value=d["frais_intermediation"])
            try:
                ws.cell(row=i, column=7, value=datetime.strptime(d["date"], "%Y-%m-%d").date())
            except ValueError:
                ws.cell(row=i, column=7, value=d["date"])
            ws.cell(row=i, column=8, value=d["mandat_total"] - d["frais_intermediation"])
        for col in (1, 3, 4, 6, 7, 8):
            ws.column_dimensions[get_column_letter(col)].width = 16
    return wb


def commerciaux_pour_export(db, entreprise_id):
    if entreprise_id and entreprise_id != "toutes":
        return db.execute(
            "SELECT * FROM commerciaux WHERE entreprise_id = ? ORDER BY nom", (entreprise_id,)
        ).fetchall()
    return db.execute("SELECT * FROM commerciaux ORDER BY nom").fetchall()


def generer_export(seulement_nouveaux):
    db = get_db()
    entreprise_id = request.args.get("entreprise_id")
    commerciaux_rows = commerciaux_pour_export(db, entreprise_id)

    dossiers_par_commercial = {}
    tous_ids = []
    for com in commerciaux_rows:
        if seulement_nouveaux:
            rows = db.execute(
                "SELECT * FROM dossiers WHERE commercial_id = ? AND exporte_excel_le IS NULL ORDER BY date DESC, id DESC",
                (com["id"],),
            ).fetchall()
        else:
            rows = db.execute(
                "SELECT * FROM dossiers WHERE commercial_id = ? ORDER BY date DESC, id DESC", (com["id"],)
            ).fetchall()
        if rows:
            dossiers_par_commercial[com["id"]] = rows
            tous_ids.extend(r["id"] for r in rows)

    if not tous_ids:
        message = "Aucun nouveau dossier depuis le dernier export" if seulement_nouveaux else "Aucun dossier à exporter"
        return jsonify(error=message), 400

    wb = construire_classeur_export(commerciaux_rows, dossiers_par_commercial)

    maintenant = datetime.utcnow().isoformat()
    db.executemany("UPDATE dossiers SET exporte_excel_le = ? WHERE id = ?", [(maintenant, i) for i in tous_ids])
    db.commit()

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    prefixe = "mise-a-jour" if seulement_nouveaux else "export"
    nom_fichier = f"{prefixe}-hse-motors-{date.today().isoformat()}.xlsx"
    return Response(
        buffer.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nom_fichier}"'},
    )


@app.post("/api/export-excel")
@admin_required
def export_excel_complet():
    return generer_export(seulement_nouveaux=False)


@app.post("/api/export-excel/maj")
@admin_required
def export_excel_maj():
    return generer_export(seulement_nouveaux=True)


# ---------- Frontend ----------

@app.get("/")
def index():
    return send_from_directory(app.static_folder, "index.html")


# Exécuté à l'import du module : garantit que la base est prête à la fois en local
# (python3 server.py) et en production (gunicorn server:app, qui n'exécute jamais __main__).
init_db()

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=8731, debug=True)
